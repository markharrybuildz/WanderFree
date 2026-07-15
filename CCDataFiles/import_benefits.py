#!/usr/bin/env python3
"""
Import card catalog data from Business_Card_Benefits.xlsx into Supabase.

Only writes the card-side catalog tables (no user data):
    card_issuers -> benefit_categories -> card_products -> benefit_definitions

The import is idempotent: every write is an upsert keyed on a unique
constraint, so re-running it updates existing rows instead of duplicating.
Rows are written in chunks so a failure partway through does not lose the
batches that already succeeded — just re-run to resume.

Usage:
    # Parse + transform + show what WOULD be written (no DB connection needed):
    python import_benefits.py --file Business_Card_Benefits.xlsx --dry-run

    # Write to Supabase (needs env vars below):
    export SUPABASE_URL="https://<project>.supabase.co"
    export SUPABASE_SERVICE_ROLE_KEY="<service-role-key>"   # bypasses RLS; server-side only
    python import_benefits.py --file Business_Card_Benefits.xlsx --apply

Options:
    --chunk-size N    rows per upsert batch (default 50)
"""

from __future__ import annotations

import argparse
import os
import sys
from collections import OrderedDict

import openpyxl

# ---------------------------------------------------------------------------
# Column mapping (see header table). Indices are 0-based into each sheet row.
# ---------------------------------------------------------------------------
# Card Benefits sheet: Network, Card Issuer, Annual Fee, Card Product,
#   Benefit Category, Benefit Name, Annual Value, Value per Use,
#   Frequency (Times/Year), Frequency/Notes, Source
CB = dict(
    network=0,
    issuer=1,
    fee=2,
    product=3,
    category=4,
    name=5,
    annual_value=6,
    value_per_use=7,
    frequency=8,
    notes=9,
    source=10,
)
# Card Summary sheet: Card Product, Network, Card Issuer, Annual Fee, ...
CS = dict(product=0, network=1, issuer=2, fee=3)


def norm_network(raw) -> str | None:
    """Map a spreadsheet network label to the card_network enum."""
    if raw is None:
        return None
    key = str(raw).strip().lower()
    return {
        "visa": "visa",
        "mastercard": "mastercard",
        "american express": "amex",
        "amex": "amex",
        "discover": "discover",
    }.get(key)


def map_frequency(freq) -> tuple[str, str, float | None]:
    """
    Translate the 'Times/Year' value into:
        (reset_frequency enum, reset_basis enum, frequency_per_year numeric)

    -1   = resets on account anniversary (annual cadence, anniversary basis)
    N/A  = applies per qualifying use, no fixed annual cap -> one_time
    0.25 = roughly every 4 years (e.g. Global Entry) -> one_time
    The raw number is preserved verbatim in frequency_per_year.
    """
    if freq is None or (isinstance(freq, str) and freq.strip().upper() == "N/A"):
        return "one_time", "calendar", None
    try:
        n = float(freq)
    except (TypeError, ValueError):
        return "annual", "calendar", None

    if n == -1:
        return "annual", "anniversary", 1.0  # once per anniversary year
    if n == 12:
        return "monthly", "calendar", n
    if n == 4:
        return "quarterly", "calendar", n
    if n == 2:
        return "semiannual", "calendar", n
    if n == 1:
        return "annual", "calendar", n
    if 0 < n < 1:
        return "one_time", "calendar", n  # 0.25 -> every ~4 years
    # 6, 8, and any other in-year recurring count: resets annually, count kept.
    return "annual", "calendar", n


def to_number(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip().replace("$", "").replace(",", "")
        if s == "" or s.upper() == "N/A":
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return float(v)


def clean_source(v) -> str | None:
    if not v:
        return None
    s = str(v).strip()
    if s.lower().startswith("source:"):
        s = s[len("source:") :].strip()
    return s or None


def is_section_divider(row, issuer_idx) -> bool:
    """Section header rows (e.g. '◆ AMERICAN EXPRESS ◆') have only the first cell."""
    return row[issuer_idx] is None


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------
def parse_workbook(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)

    issuers: "OrderedDict[str, None]" = OrderedDict()
    categories: "OrderedDict[str, None]" = OrderedDict()
    products: "OrderedDict[tuple[str, str], dict]" = (
        OrderedDict()
    )  # (issuer, name) -> row
    benefits: list[dict] = []

    def add_product(issuer, name, network, fee):
        issuer = str(issuer).strip()
        name = str(name).strip()
        issuers.setdefault(issuer, None)
        key = (issuer, name)
        if key not in products:
            products[key] = {
                "issuer": issuer,
                "name": name,
                "network": norm_network(network),
                "annual_fee": to_number(fee) or 0,
            }

    # --- Card Benefits sheet (skip title, covers, header -> rows[3:]) ---
    cb = wb["Card Benefits"]
    for row in list(cb.iter_rows(values_only=True))[3:]:
        if is_section_divider(row, CB["issuer"]):
            continue
        issuer = str(row[CB["issuer"]]).strip()
        product = str(row[CB["product"]]).strip()
        category = str(row[CB["category"]]).strip()
        add_product(issuer, product, row[CB["network"]], row[CB["fee"]])
        categories.setdefault(category, None)

        freq_enum, basis, freq_py = map_frequency(row[CB["frequency"]])
        notes = row[CB["notes"]]
        benefits.append(
            {
                "issuer": issuer,
                "product": product,
                "category": category,
                "name": str(row[CB["name"]]).strip(),
                "annual_value": to_number(row[CB["annual_value"]]),
                "value_per_period": to_number(row[CB["value_per_use"]]),
                "reset_frequency": freq_enum,
                "reset_basis": basis,
                "frequency_per_year": freq_py,
                "description": notes,
                "requires_enrollment": bool(notes and "enroll" in str(notes).lower()),
                "source_url": clean_source(row[CB["source"]]),
            }
        )

    # --- Card Summary sheet (skip title, header -> rows[2:]) -> extra products ---
    cs = wb["Card Summary"]
    for row in list(cs.iter_rows(values_only=True))[2:]:
        if is_section_divider(row, CS["network"]):  # dividers have no network
            continue
        add_product(
            row[CS["issuer"]], row[CS["product"]], row[CS["network"]], row[CS["fee"]]
        )

    return {
        "issuers": list(issuers.keys()),
        "categories": list(categories.keys()),
        "products": list(products.values()),
        "benefits": benefits,
    }


# ---------------------------------------------------------------------------
# Loading (chunked, idempotent upserts)
# ---------------------------------------------------------------------------
def chunked(seq, size):
    for i in range(0, len(seq), size):
        yield seq[i : i + size]


def upsert_chunks(client, table, rows, on_conflict, chunk_size, key_fn):
    """Upsert rows in batches; return {key_fn(returned_row): id}. Fail-soft per chunk."""
    id_map: dict = {}
    failed = 0
    total = len(rows)
    for n, batch in enumerate(chunked(rows, chunk_size), 1):
        try:
            resp = (
                client.table(table)
                .upsert(batch, on_conflict=on_conflict, returning="representation")
                .execute()
            )
            for r in resp.data:
                id_map[key_fn(r)] = r["id"]
            done = min(n * chunk_size, total)
            print(f"    {table}: {done}/{total} upserted")
        except Exception as e:  # noqa: BLE001 — keep going so prior chunks persist
            failed += len(batch)
            print(
                f"    !! {table} chunk {n} failed ({len(batch)} rows): {e}",
                file=sys.stderr,
            )
    if failed:
        print(
            f"    {table}: {failed} rows failed — safe to re-run to retry.",
            file=sys.stderr,
        )
    return id_map


def apply_to_supabase(parsed, chunk_size):
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get(
        "SUPABASE_SERVICE_KEY"
    )
    if not url or not key:
        sys.exit("ERROR: set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY to apply.")

    try:
        from supabase import create_client
    except ImportError:
        sys.exit("ERROR: pip install supabase  (needed only for --apply)")

    client = create_client(url, key)

    print("[1/4] card_issuers")
    issuer_ids = upsert_chunks(
        client,
        "card_issuers",
        [{"name": n} for n in parsed["issuers"]],
        on_conflict="name",
        chunk_size=chunk_size,
        key_fn=lambda r: r["name"],
    )

    print("[2/4] benefit_categories")
    category_ids = upsert_chunks(
        client,
        "benefit_categories",
        [{"name": n} for n in parsed["categories"]],
        on_conflict="name",
        chunk_size=chunk_size,
        key_fn=lambda r: r["name"],
    )

    print("[3/4] card_products")
    product_rows = []
    for p in parsed["products"]:
        iid = issuer_ids.get(p["issuer"])
        if iid is None:
            print(
                f"    !! no issuer id for {p['issuer']!r}; skipping {p['name']!r}",
                file=sys.stderr,
            )
            continue
        product_rows.append(
            {
                "issuer_id": iid,
                "name": p["name"],
                "network": p["network"],
                "annual_fee": p["annual_fee"],
            }
        )
    product_ids = upsert_chunks(
        client,
        "card_products",
        product_rows,
        on_conflict="issuer_id,name",
        chunk_size=chunk_size,
        key_fn=lambda r: (r["issuer_id"], r["name"]),
    )
    # also key products by (issuer_name, product_name) for benefit linking
    prod_by_name = {}
    for p in parsed["products"]:
        iid = issuer_ids.get(p["issuer"])
        pid = product_ids.get((iid, p["name"]))
        if pid:
            prod_by_name[(p["issuer"], p["name"])] = pid

    print("[4/4] benefit_definitions")
    benefit_rows = []
    for b in parsed["benefits"]:
        pid = prod_by_name.get((b["issuer"], b["product"]))
        cid = category_ids.get(b["category"])
        if pid is None:
            print(
                f"    !! no product id for {b['product']!r}; skipping benefit "
                f"{b['name']!r}",
                file=sys.stderr,
            )
            continue
        benefit_rows.append(
            {
                "card_product_id": pid,
                "benefit_category_id": cid,
                "name": b["name"],
                "description": b["description"],
                "value_per_period": b["value_per_period"],
                "annual_value": b["annual_value"],
                "reset_frequency": b["reset_frequency"],
                "reset_basis": b["reset_basis"],
                "frequency_per_year": b["frequency_per_year"],
                "requires_enrollment": b["requires_enrollment"],
                "source_url": b["source_url"],
            }
        )
    upsert_chunks(
        client,
        "benefit_definitions",
        benefit_rows,
        on_conflict="card_product_id,name",
        chunk_size=chunk_size,
        key_fn=lambda r: (r["card_product_id"], r["name"]),
    )
    print("Done.")


# ---------------------------------------------------------------------------
def print_dry_run(parsed, chunk_size):
    from collections import Counter

    b = parsed["benefits"]
    print("DRY RUN — nothing written.\n")
    print(f"  card_issuers        : {len(parsed['issuers'])}")
    print(f"  benefit_categories  : {len(parsed['categories'])}")
    print(f"  card_products       : {len(parsed['products'])}")
    print(f"  benefit_definitions : {len(b)}")
    print()
    print(
        "  network distribution :",
        dict(Counter(p["network"] for p in parsed["products"])),
    )
    print("  reset_frequency      :", dict(Counter(x["reset_frequency"] for x in b)))
    print("  reset_basis          :", dict(Counter(x["reset_basis"] for x in b)))
    print(
        "  requires_enrollment  :", dict(Counter(x["requires_enrollment"] for x in b))
    )
    print("  frequency_per_year   :", dict(Counter(x["frequency_per_year"] for x in b)))
    print()
    chunks = (len(b) + chunk_size - 1) // chunk_size
    print(f"  benefit_definitions would upload in {chunks} chunk(s) of {chunk_size}.")
    print("\n  sample benefit row:")
    for k, v in b[0].items():
        sv = (str(v)[:70] + "…") if v and len(str(v)) > 70 else v
        print(f"    {k:20}: {sv}")


def main():
    ap = argparse.ArgumentParser(description="Import card catalog into Supabase.")
    ap.add_argument("--file", default="Business_Card_Benefits.xlsx")
    ap.add_argument("--chunk-size", type=int, default=50)
    g = ap.add_mutually_exclusive_group()
    g.add_argument(
        "--dry-run", action="store_true", help="parse + preview, no writes (default)"
    )
    g.add_argument("--apply", action="store_true", help="write to Supabase")
    args = ap.parse_args()

    parsed = parse_workbook(args.file)
    if args.apply:
        apply_to_supabase(parsed, args.chunk_size)
    else:
        print_dry_run(parsed, args.chunk_size)


if __name__ == "__main__":
    main()
